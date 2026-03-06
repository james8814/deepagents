"""Excel file converter using openpyxl."""

import logging
import time
from pathlib import Path

from deepagents.middleware.converters.base import BaseConverter

logger = logging.getLogger(__name__)


class XLSXConverter(BaseConverter):
    """Converter for Excel files to Markdown tables.

    Supports .xlsx and .xls files with multiple sheets.

    Dependencies:
        openpyxl: Install with `pip install openpyxl`
    """

    def convert(self, path: Path, raw_content: str | bytes | None = None) -> str:
        """Convert an Excel file to Markdown.

        Args:
            path: Path to the Excel file.
            raw_content: Ignored for Excel files (binary format).

        Returns:
            Markdown-formatted string with all sheets as tables.
        """
        import openpyxl

        start = time.time()

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        parts = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            parts.append(f"## Sheet: {sheet_name}\n")

            # Extract data
            rows = []
            for row in sheet.iter_rows(values_only=True):
                # Convert all values to strings
                row_values = [str(cell) if cell is not None else "" for cell in row]
                # Skip completely empty rows
                if any(v for v in row_values):
                    rows.append(row_values)

            if rows:
                # Limit output for large sheets
                max_rows = 100
                if len(rows) > max_rows:
                    display_rows = rows[:max_rows]
                    truncated = True
                else:
                    display_rows = rows
                    truncated = False

                # Use first row as headers
                headers = display_rows[0] if display_rows else []
                data_rows = display_rows[1:] if len(display_rows) > 1 else []

                parts.append(self._format_as_table(data_rows, headers))

                if truncated:
                    parts.append(f"\n*... ({len(rows) - max_rows} more rows truncated)*")
            else:
                parts.append("(Empty sheet)")

        wb.close()

        duration = time.time() - start
        self._log_conversion(path, duration)

        return "\n\n".join(parts)
